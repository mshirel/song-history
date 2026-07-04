"""FastAPI + HTMX web UI for the worship catalog."""

from __future__ import annotations

import base64
import csv
import importlib.metadata
import io
import json
import logging
import math
import os
import platform
import re
import secrets
import threading
import time
from collections import defaultdict
from collections.abc import AsyncGenerator
from concurrent.futures import ThreadPoolExecutor
from contextlib import asynccontextmanager
from datetime import date, datetime
from pathlib import Path
from typing import Any

from fastapi import (
    Depends,
    FastAPI,
    Form,
    HTTPException,
    Query,
    Request,
    Response,
    UploadFile,
)
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from prometheus_client import CONTENT_TYPE_LATEST, REGISTRY, Counter, Histogram, generate_latest
from starlette.exceptions import HTTPException as StarletteHTTPException
from starlette.middleware.base import BaseHTTPMiddleware, RequestResponseEndpoint

from worship_catalog.db import Database
from worship_catalog.import_service import run_import
from worship_catalog.log_config import RequestLoggingMiddleware
from worship_catalog.log_config import setup as _setup_logging
from worship_catalog.notify import send_pushover
from worship_catalog.service_slots import (
    DEFAULT_WINDOW_DAYS,
    VALID_SLOTS,
    WINDOW_OPTIONS,
)
from worship_catalog.services.missing_services_report import compute_missing_services
from worship_catalog.services.report_service import compute_stats_data
from worship_catalog.web.csrf import SelfHealingCSRFMiddleware

_setup_logging()
_log = logging.getLogger("worship_catalog.web")


# Bounded thread pool for background import jobs (#52)
# Declared here (before lifespan) so the lifespan can shut it down gracefully (#135).
#
# Shutdown timeline (#295):
#   SIGTERM → uvicorn stops accepting connections → lifespan __aexit__ fires
#   → _shutdown_executor waits up to _EXECUTOR_SHUTDOWN_TIMEOUT seconds
#   → compose stop_grace_period (35s) > timeout (30s), so SIGKILL comes after
#   → SQLite WAL mode recovers automatically if SIGKILL hits mid-write
#   Requires init:true in compose.yml (#289) so SIGTERM reaches uvicorn.
_MAX_IMPORT_WORKERS: int = 4
_EXECUTOR_SHUTDOWN_TIMEOUT: int = 30  # seconds to wait for in-flight jobs before giving up
_import_executor = ThreadPoolExecutor(max_workers=_MAX_IMPORT_WORKERS)


@asynccontextmanager
async def _lifespan(app: FastAPI) -> AsyncGenerator[None, None]:
    """Run startup tasks before the app begins serving requests."""
    db = _get_db()
    try:
        db.purge_old_import_jobs(days=90)
        _log.info("Startup purge: removed import_jobs older than 90 days")
    except Exception as exc:  # noqa: BLE001
        _log.warning("Startup purge failed", extra={"error": str(exc)})
    finally:
        db.close()
    try:
        yield
    finally:
        # Graceful shutdown: give in-flight import jobs time to finish (#135).
        # Use a background thread so we can enforce a hard timeout without
        # blocking the event loop indefinitely.
        shutdown_done = threading.Event()

        def _shutdown_executor() -> None:
            _import_executor.shutdown(wait=True, cancel_futures=False)
            shutdown_done.set()

        t = threading.Thread(target=_shutdown_executor, daemon=True)
        t.start()
        if not shutdown_done.wait(timeout=_EXECUTOR_SHUTDOWN_TIMEOUT):
            _log.warning(
                "Executor did not finish within timeout — proceeding with shutdown",
                extra={"timeout_seconds": _EXECUTOR_SHUTDOWN_TIMEOUT},
            )


app = FastAPI(title="Worship Catalog", lifespan=_lifespan)

# CSRF protection — must be added BEFORE RequestLoggingMiddleware so that 403
# responses are logged correctly. A stable CSRF_SECRET is REQUIRED in
# production: an ephemeral per-process secret invalidates every outstanding
# token on restart and fails CSRF validation across workers (#406). Outside
# TESTING mode a missing secret is a hard error so a misconfigured deploy fails
# fast instead of silently degrading. TESTING=1 keeps an ephemeral secret for
# the test suite / local dev.
_TESTING: bool = os.environ.get("TESTING", "").strip() in ("1", "true", "yes")
_csrf_secret_env = os.environ.get("CSRF_SECRET")
if _csrf_secret_env:
    _CSRF_SECRET = _csrf_secret_env
elif _TESTING:
    _CSRF_SECRET = secrets.token_hex(32)
    _log.warning(
        "CSRF_SECRET not set — using an ephemeral random secret (TESTING mode). "
        "Set a stable CSRF_SECRET for any real deployment."
    )
else:
    raise RuntimeError(
        "CSRF_SECRET is not set. Generate a stable secret once with "
        '`python3 -c "import secrets; print(secrets.token_hex(32))"` and set it '
        "in the environment so CSRF tokens survive restarts and work across "
        "workers. For tests or local dev, set TESTING=1 to allow an ephemeral "
        "random secret."
    )
# cookie_name is set explicitly so the coupling with client-side JS
# (upload.js, reports.js) and CsrfAwareClient in conftest.py is visible (#239).
app.add_middleware(
    SelfHealingCSRFMiddleware,
    secret=_CSRF_SECRET,
    cookie_name="csrftoken",
    exempt_urls=[re.compile(r"^/health$"), re.compile(r"^/metrics$")],
)
app.add_middleware(RequestLoggingMiddleware)

# Content-Security-Policy — defence-in-depth against XSS (#197).
# All scripts must be external files served from /static/ (no inline JS).
_CSP_POLICY: str = (
    "default-src 'self'; "
    "script-src 'self'; "
    "style-src 'self' 'unsafe-inline'; "
    "img-src 'self' data:; "
    "frame-ancestors 'none'"
)


# HSTS is opt-in via HTTPS_ONLY so that HTTP-only local/CI deployments don't
# send a policy that would lock a browser into HTTPS for the host (#405).
_HSTS_ENABLED: bool = os.environ.get("HTTPS_ONLY", "").strip() in ("1", "true", "yes")
_HSTS_VALUE: str = "max-age=31536000; includeSubDomains"


class _SecurityHeadersMiddleware(BaseHTTPMiddleware):
    """Attach security headers to every response (#197, #282, #405)."""

    async def dispatch(self, request: Request, call_next: RequestResponseEndpoint) -> Response:
        response = await call_next(request)
        response.headers["Content-Security-Policy"] = _CSP_POLICY
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Permissions-Policy"] = (
            "camera=(), microphone=(), geolocation=(), payment=()"
        )
        # Only assert HSTS when the deployment is HTTPS-only (behind Cloudflare/
        # Traefik). Sending it over plain HTTP would wrongly pin clients to HTTPS.
        if _HSTS_ENABLED:
            response.headers["Strict-Transport-Security"] = _HSTS_VALUE
        return response


app.add_middleware(_SecurityHeadersMiddleware)

try:
    _HTTP_REQUESTS_TOTAL: Counter = Counter(
        "http_requests_total",
        "Total HTTP requests counted by method, path, and status code",
        ["method", "path", "status_code"],
    )
except ValueError:
    # Module was reloaded (e.g., during tests via importlib.reload); reuse the
    # existing collector rather than double-registering in the global registry.
    _HTTP_REQUESTS_TOTAL = REGISTRY._names_to_collectors[  # type: ignore[assignment]
        "http_requests_total"
    ]

try:
    # Request-latency histogram so Grafana/Prometheus can compute p50/p95/p99
    # per route (#395). Default buckets suit sub-second web responses.
    _HTTP_REQUEST_DURATION_SECONDS: Histogram = Histogram(
        "http_request_duration_seconds",
        "HTTP request duration in seconds by method and path",
        ["method", "path"],
    )
except ValueError:
    _HTTP_REQUEST_DURATION_SECONDS = REGISTRY._names_to_collectors[  # type: ignore[assignment]
        "http_request_duration_seconds"
    ]


class _PrometheusMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next: RequestResponseEndpoint) -> Response:
        start = time.perf_counter()
        response = await call_next(request)
        if request.url.path != "/metrics":
            elapsed = time.perf_counter() - start
            _HTTP_REQUESTS_TOTAL.labels(
                method=request.method,
                path=request.url.path,
                status_code=str(response.status_code),
            ).inc()
            _HTTP_REQUEST_DURATION_SECONDS.labels(
                method=request.method,
                path=request.url.path,
            ).observe(elapsed)
        return response


app.add_middleware(_PrometheusMiddleware)

_TEMPLATES_DIR = Path(__file__).parent / "templates"
_STATIC_DIR = Path(__file__).parent / "static"
templates = Jinja2Templates(directory=str(_TEMPLATES_DIR))
app.mount("/static", StaticFiles(directory=str(_STATIC_DIR)), name="static")


# Headers added by the public proxy chain (Cloudflare tunnel → Traefik). Their
# presence means the request did NOT come from the internal Prometheus scrape,
# which connects directly to the host-published :8000 with none of these (#449).
_PROXY_HEADERS: tuple[str, ...] = (
    "x-forwarded-for",
    "x-forwarded-host",
    "x-forwarded-proto",
    "cf-connecting-ip",
    "forwarded",
)


@app.get("/metrics", include_in_schema=False)
async def metrics(request: Request) -> Response:
    """Prometheus metrics — internal scraping only (#449).

    Public requests reach the app through the proxy chain (Cloudflare tunnel →
    Traefik) and carry forwarding headers (X-Forwarded-For etc.); the internal
    Prometheus scrape hits :8000 directly with none of them. Deny the proxied
    path with 404 so the route map / traffic / latency aren't exposed publicly,
    while the internal scrape keeps working.
    """
    if any(h in request.headers for h in _PROXY_HEADERS):
        raise HTTPException(status_code=404)
    return Response(content=generate_latest(), media_type=CONTENT_TYPE_LATEST)


def _is_htmx_request(request: Request) -> bool:
    """True when the request came from HTMX (it sets the ``HX-Request`` header)."""
    return request.headers.get("HX-Request", "").lower() == "true"


@app.exception_handler(StarletteHTTPException)
async def http_exception_handler(request: Request, exc: StarletteHTTPException) -> HTMLResponse:
    # HTMX only swaps 2xx responses and expects a fragment, not a whole page.
    # Return a small inline error partial for HTMX callers so a 422 date-range
    # error surfaces in the result region instead of being silently swallowed —
    # reports.js drops the fragment into the form's target (#496).
    if _is_htmx_request(request):
        return templates.TemplateResponse(
            request, "_error.html", {"detail": str(exc.detail)},
            status_code=exc.status_code,
            headers=getattr(exc, "headers", None),
        )
    if exc.status_code == 404:
        return templates.TemplateResponse(
            request, "404.html", {"detail": str(exc.detail)}, status_code=404
        )
    return templates.TemplateResponse(
        request, "500.html", {"detail": str(exc.detail)},
        status_code=exc.status_code,
        headers=getattr(exc, "headers", None),
    )


@app.exception_handler(Exception)
async def unhandled_exception_handler(request: Request, exc: Exception) -> HTMLResponse:
    _log.exception("Unhandled exception", extra={"path": str(request.url)})
    detail = "An unexpected error occurred."
    if _is_htmx_request(request):
        return templates.TemplateResponse(
            request, "_error.html", {"detail": detail}, status_code=500
        )
    return templates.TemplateResponse(
        request, "500.html", {"detail": detail}, status_code=500
    )


_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")

# Whitelist pattern for safe filename characters in HTTP headers and filesystems.
# Keeps only alphanumeric characters, hyphens, underscores, and dots.
# Spaces are replaced so that Content-Disposition filenames are unambiguously
# safe without relying on RFC 6266 quoted-string parsing by every client.
# CR and LF are always excluded to prevent HTTP header injection.
_SAFE_FILENAME_RE = re.compile(r"[^\w.\-]|[\r\n]")


def _sanitize_header_filename(raw: str) -> str:
    """Strip unsafe characters from a filename used in Content-Disposition headers.

    Uses Path.name to strip directory components, then replaces any character
    that is not alphanumeric, a hyphen, underscore, or dot with an underscore.
    Spaces are replaced with underscores for maximum client compatibility.
    CR and LF are always replaced to prevent HTTP header injection.
    """
    basename = Path(raw).name
    return _SAFE_FILENAME_RE.sub("_", basename)

# Docker-baked version/build metadata (#261, #262)
_VERSION_FILE: Path = Path("/app/.version")
_BUILD_DATE_FILE: Path = Path("/app/.build-date")

def _format_build_date(raw: str) -> str:
    """Format an ISO-8601 build timestamp to human-readable local time (#331)."""
    if not raw or raw == "development":
        return raw
    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        local_dt = dt.astimezone()  # convert to server local timezone
        return local_dt.strftime("%B %-d, %Y at %-I:%M %p %Z")
    except (ValueError, OSError):
        return raw


# Upload constants (#45)
MAX_UPLOAD_BYTES: int = 200 * 1024 * 1024  # 200 MB
_UPLOAD_CHUNK_SIZE: int = 64 * 1024  # 64 KB — chunk size for streaming upload reads (#297)
_PPTX_MIME = (
    "application/vnd.openxmlformats-officedocument.presentationml.presentation"
)

# Per-client upload rate limiting (#173)
_UPLOAD_RATE_LIMIT: int = 10  # max uploads per window
_UPLOAD_RATE_WINDOW_SECONDS: int = 3600  # 1 hour
# Env-configurable trusted proxy support (#283, #404).  When TRUSTED_PROXY=1 the
# rate limiter identifies clients via Cloudflare's unspoofable CF-Connecting-IP
# (or, failing that, the proxy-appended rightmost X-Forwarded-For entry) instead
# of the socket peer, so clients behind the proxy are bucketed individually.
_TRUST_PROXY: bool = os.environ.get("TRUSTED_PROXY", "").strip() in ("1", "true", "yes")


def _get_client_ip(request: Request) -> str:
    """Return the real client IP, respecting trusted-proxy headers (#283, #404).

    The leftmost ``X-Forwarded-For`` entry is supplied by the client and is
    therefore spoofable — trusting it lets an attacker mint a fresh rate-limit
    bucket per request. This deployment sits behind Cloudflare, which sets the
    unspoofable ``CF-Connecting-IP`` header (and strips any client-supplied
    copy), so prefer it. If it is absent, fall back to the *rightmost*
    ``X-Forwarded-For`` entry — the address appended by our own proxy — rather
    than the client-controlled leftmost one.
    """
    if _TRUST_PROXY:
        cf_ip = request.headers.get("cf-connecting-ip", "").strip()
        if cf_ip:
            return cf_ip
        xff = request.headers.get("x-forwarded-for", "")
        if xff:
            parts = [p.strip() for p in xff.split(",") if p.strip()]
            if parts:
                return parts[-1]
    return request.client.host if request.client else "unknown"


class _UploadRateLimiter:
    """Thread-safe sliding-window rate limiter keyed by client IP.

    When *db_path* is provided, timestamps are persisted to a SQLite table
    (``rate_limit_events``) so that rate-limit state survives process restarts
    (#241).  Without a *db_path* the limiter falls back to an in-memory dict
    (useful for tests that don't need persistence).
    """

    def __init__(
        self,
        db_path: Path | None = None,
        limit: int | None = None,
        window_seconds: int | None = None,
        table: str = "rate_limit_events",
    ) -> None:
        self._lock = threading.Lock()
        self._db_path = db_path
        # When no explicit override is given, fall back to the module constants
        # at CALL time (see _limit/_window) so tests that patch them still work.
        self._limit_override = limit
        self._window_override = window_seconds
        # Table name is an internal literal (never user input) — safe to inline.
        self._table = table
        self._conn: Any = None  # persistent SQLite connection (#341)
        # In-memory fallback when no db_path is given
        self._timestamps: dict[str, list[float]] = defaultdict(list)
        if db_path is not None:
            self._init_db()

    @property
    def _limit(self) -> int:
        return self._limit_override if self._limit_override is not None else _UPLOAD_RATE_LIMIT

    @property
    def _window(self) -> int:
        if self._window_override is not None:
            return self._window_override
        return _UPLOAD_RATE_WINDOW_SECONDS

    # -- private helpers --------------------------------------------------

    def _init_db(self) -> None:
        import sqlite3

        self._conn = sqlite3.connect(self._db_path, check_same_thread=False)  # type: ignore[arg-type]
        self._conn.execute(
            f"CREATE TABLE IF NOT EXISTS {self._table} "
            "(client_ip TEXT NOT NULL, timestamp REAL NOT NULL)"
        )
        self._conn.execute(
            f"CREATE INDEX IF NOT EXISTS idx_{self._table}_ip_ts "
            f"ON {self._table} (client_ip, timestamp)"
        )
        self._conn.commit()

    def _db_is_allowed(self, client_ip: str) -> tuple[bool, int]:
        now = time.time()
        window_start = now - self._window

        conn = self._conn
        # Prune expired entries for this IP
        conn.execute(
            f"DELETE FROM {self._table} WHERE client_ip = ? AND timestamp <= ?",
            (client_ip, window_start),
        )
        row = conn.execute(
            f"SELECT COUNT(*) FROM {self._table} WHERE client_ip = ? AND timestamp > ?",
            (client_ip, window_start),
        ).fetchone()
        count = row[0] if row else 0

        if count >= self._limit:
            oldest_row = conn.execute(
                f"SELECT MIN(timestamp) FROM {self._table} "
                "WHERE client_ip = ? AND timestamp > ?",
                (client_ip, window_start),
            ).fetchone()
            oldest_ts = oldest_row[0] if oldest_row and oldest_row[0] else now
            retry_after = int(oldest_ts - window_start) + 1
            conn.commit()
            return False, max(retry_after, 1)

        conn.execute(
            f"INSERT INTO {self._table} (client_ip, timestamp) VALUES (?, ?)",
            (client_ip, now),
        )
        conn.commit()
        return True, 0

    def _mem_is_allowed(self, client_ip: str) -> tuple[bool, int]:
        now = time.monotonic()
        window_start = now - self._window
        timestamps = self._timestamps[client_ip]
        self._timestamps[client_ip] = [
            t for t in timestamps if t > window_start
        ]
        timestamps = self._timestamps[client_ip]
        if len(timestamps) >= self._limit:
            oldest_in_window = timestamps[0]
            retry_after = int(oldest_in_window - window_start) + 1
            return False, max(retry_after, 1)
        timestamps.append(now)
        return True, 0

    # -- public API -------------------------------------------------------

    def is_allowed(self, client_ip: str) -> tuple[bool, int]:
        """Check whether the client is within its rate budget.

        Returns (allowed, retry_after_seconds).
        """
        with self._lock:
            if self._db_path is not None:
                return self._db_is_allowed(client_ip)
            return self._mem_is_allowed(client_ip)


# Report endpoints are public and run unbounded aggregation / openpyxl builds,
# so they get their own (more generous than upload) per-client budget (#450).
_REPORT_RATE_LIMIT: int = 60
_REPORT_RATE_WINDOW_SECONDS: int = 300  # 5 minutes


def _limiter_db_path() -> Path:
    """Shared SQLite file (next to the app DB) backing all rate limiters."""
    return Path(os.environ.get("DB_PATH", "data/worship.db")).parent / "rate_limits.db"


# Upload limiter: no explicit limit/window → reads _UPLOAD_RATE_LIMIT/_WINDOW at
# call time (preserves the existing patch-the-constant test behaviour).
_upload_limiter = _UploadRateLimiter(db_path=_limiter_db_path(), table="rate_limit_events")
# Report limiter: its own, more generous budget on a separate table.
_report_limiter = _UploadRateLimiter(
    db_path=_limiter_db_path(),
    limit=_REPORT_RATE_LIMIT,
    window_seconds=_REPORT_RATE_WINDOW_SECONDS,
    table="report_rate_limit_events",
)


def _report_rate_limit(request: Request) -> None:
    """FastAPI dependency: throttle public report generation per client (#450)."""
    allowed, retry_after = _report_limiter.is_allowed(_get_client_ip(request))
    if not allowed:
        _log.warning(
            "Report rate limit exceeded",
            extra={"client_ip": _get_client_ip(request), "retry_after": retry_after},
        )
        raise HTTPException(
            status_code=429,
            detail="Report rate limit exceeded — try again later",
            headers={"Retry-After": str(retry_after)},
        )


def _validate_date_range(start_date: str, end_date: str) -> None:
    """Raise HTTPException 422 if dates are not valid ISO-8601 or range is inverted."""
    for label, val in (("start_date", start_date), ("end_date", end_date)):
        if not _DATE_RE.match(val):
            raise HTTPException(
                status_code=422,
                detail=f"Invalid {label}: '{val}' — expected YYYY-MM-DD",
            )
        try:
            date.fromisoformat(val)
        except ValueError as exc:
            raise HTTPException(
                status_code=422,
                detail=f"Invalid {label}: '{val}' — not a real calendar date",
            ) from exc
    if start_date > end_date:
        raise HTTPException(
            status_code=422,
            detail=f"start_date ({start_date}) must not be after end_date ({end_date})",
        )


_schema_ready: bool = False
_schema_lock = threading.Lock()


def _get_db() -> Database:
    global _schema_ready  # noqa: PLW0603
    db_path = Path(os.environ.get("DB_PATH", "data/worship.db"))
    db = Database(db_path)
    db.connect()
    if not _schema_ready:
        with _schema_lock:
            if not _schema_ready:  # double-check under lock (#277, #296)
                db.init_schema()
                _schema_ready = True
    return db


async def get_db() -> AsyncGenerator[Database, None]:
    """FastAPI dependency that always closes the DB connection (#236).

    Async generator so that FastAPI runs it in the event-loop thread — the
    same thread used by ``async def`` route handlers.  A sync generator would
    be dispatched to the default threadpool, causing SQLite's
    ``check_same_thread`` assertion to fail.
    """
    db = _get_db()
    try:
        yield db
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------


@app.get("/health")
async def health(response: Response) -> dict[str, str]:
    """Return 200 if DB is reachable; 503 otherwise (issue #31).

    Uses manual _get_db()/close() instead of Depends(get_db) because
    the health check must return 503 (not 500) when the DB is unreachable,
    including when _get_db() itself raises.
    """
    try:
        db = _get_db()
        db.cursor().execute("SELECT 1")
        db.close()
        return {"status": "ok"}
    except Exception as exc:
        _log.warning("Health check DB failure", extra={"error": str(exc)})
        response.status_code = 503
        return {"status": "error"}


@app.get("/", response_class=RedirectResponse)
async def root() -> RedirectResponse:
    return RedirectResponse(url="/songs")


def _is_htmx_history_restore(request: Request) -> bool:
    """True when HTMX is replaying browser history and needs the full page shell."""
    return request.headers.get("HX-History-Restore-Request", "").lower() == "true"


@app.get("/songs", response_class=HTMLResponse)
async def songs(
    request: Request,
    q: str | None = Query(default=None),
    sort: str = Query(default="performance_count"),
    sort_dir: str = Query(default="desc"),
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=50, ge=10, le=500),
    db: Database = Depends(get_db),  # noqa: B008
) -> HTMLResponse:
    sort = sort if sort in Database._SONGS_SORT_COLS else "performance_count"
    sort_dir = "asc" if sort_dir == "asc" else "desc"
    rows, total = db.query_songs_paginated(
        q, sort=sort, sort_dir=sort_dir, page=page, per_page=per_page,
    )
    library_size = db.count_songs()

    total_pages = math.ceil(total / per_page) if total > 0 else 1

    context = {
        "songs": rows, "q": q or "", "sort": sort, "sort_dir": sort_dir,
        "page": page, "per_page": per_page, "total_pages": total_pages, "total": total,
        "library_size": library_size,
    }
    # HTMX search/sort swaps the whole #songs-results region (table + footer)
    # so the pagination footer never drifts from the rows shown (#386).
    template = (
        "songs.html"
        if _is_htmx_history_restore(request) or not request.headers.get("HX-Request")
        else "_songs_results.html"
    )
    return templates.TemplateResponse(request, template, context)


@app.get("/reports", response_class=HTMLResponse)
async def reports_page(
    request: Request,
    db: Database = Depends(get_db),  # noqa: B008
) -> HTMLResponse:
    return templates.TemplateResponse(
        request,
        "reports.html",
        {
            "leaders": db.query_distinct_song_leaders(),
            "window_options": WINDOW_OPTIONS,
            "default_window": str(DEFAULT_WINDOW_DAYS),
        },
    )


@app.get("/about", response_class=HTMLResponse)
async def about_page(request: Request) -> HTMLResponse:
    """Render the About page with app purpose, version, and copyright (#232, #261, #262)."""
    # Prefer baked-in file from Docker build; fall back to importlib.metadata (#261)
    if _VERSION_FILE.is_file():
        version = _VERSION_FILE.read_text().strip()
    else:
        try:
            version = importlib.metadata.version("worship-catalog")
        except importlib.metadata.PackageNotFoundError:
            version = "development"
    python_version = platform.python_version()
    db_path = Path(os.environ.get("DB_PATH", "data/worship.db")).name
    # Prefer baked-in file from Docker build; fall back to "development" (#262)
    if _BUILD_DATE_FILE.is_file():
        build_date = _format_build_date(_BUILD_DATE_FILE.read_text().strip())
    else:
        build_date = "development"
    return templates.TemplateResponse(
        request,
        "about.html",
        {
            "version": version,
            "python_version": python_version,
            "db_path": db_path,
            "build_date": build_date,
        },
    )


def _compute_stats(
    db: Database,
    start_date: str,
    end_date: str,
    leader: str,
    all_songs: bool,
) -> dict[str, Any]:
    """Thin wrapper — delegates to services.report_service.compute_stats_data (#25)."""
    return compute_stats_data(db, start_date, end_date, leader or None, all_songs)


@app.post("/reports/stats", response_class=HTMLResponse)
async def reports_stats(
    request: Request,
    start_date: str = Form(...),
    end_date: str = Form(...),
    leader: str = Form(default=""),
    all_songs: bool = Form(default=False),
    db: Database = Depends(get_db),  # noqa: B008
    _rl: None = Depends(_report_rate_limit),  # noqa: B008
) -> HTMLResponse:
    _validate_date_range(start_date, end_date)
    data = _compute_stats(db, start_date, end_date, leader, all_songs)

    _log.info(
        "Stats report generated",
        extra={"start_date": start_date, "end_date": end_date, "leader": leader or None,
               "services": len(data["services"]), "unique_songs": len(data["sorted_songs"])},
    )
    return templates.TemplateResponse(
        request,
        "stats_result.html",
        {
            "start_date": start_date,
            "end_date": end_date,
            "leader": leader,
            "all_songs": all_songs,
            **data,
        },
    )


@app.post("/reports/stats/csv")
async def reports_stats_csv(
    start_date: str = Form(...),
    end_date: str = Form(...),
    leader: str = Form(default=""),
    all_songs: bool = Form(default=False),
    db: Database = Depends(get_db),  # noqa: B008
    _rl: None = Depends(_report_rate_limit),  # noqa: B008
) -> StreamingResponse:
    _validate_date_range(start_date, end_date)
    data = _compute_stats(db, start_date, end_date, leader, all_songs)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Rank", "Title", "Credits", "Count"])
    for rank, (title, count) in enumerate(data["sorted_songs"], 1):
        writer.writerow([rank, title, data["song_credits"].get(title, ""), count])

    output.seek(0)
    filename = _sanitize_header_filename(f"stats_{start_date}_{end_date}.csv")
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/reports/stats/xlsx")
async def reports_stats_xlsx(
    start_date: str = Form(...),
    end_date: str = Form(...),
    leader: str = Form(default=""),
    all_songs: bool = Form(default=False),
    db: Database = Depends(get_db),  # noqa: B008
    _rl: None = Depends(_report_rate_limit),  # noqa: B008
) -> StreamingResponse:
    _validate_date_range(start_date, end_date)
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError as exc:
        raise HTTPException(
            status_code=501,
            detail="Excel export requires openpyxl. Install with: pip install openpyxl",
        ) from exc

    data = _compute_stats(db, start_date, end_date, leader, all_songs)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Top Songs"
    header = ["Rank", "Title", "Credits", "Count"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for rank, (title, count) in enumerate(data["sorted_songs"], 1):
        ws.append([rank, title, data["song_credits"].get(title, ""), count])

    if data["leader_breakdown"]:
        ws2 = wb.create_sheet("By Leader")
        ws2.append(["Leader", "Song", "Count"])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
        for ldr, songs in data["leader_breakdown"].items():
            for title, count in songs:
                ws2.append([ldr, title, count])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = _sanitize_header_filename(f"stats_{start_date}_{end_date}.xlsx")
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/reports/ccli/preview", response_class=HTMLResponse)
async def reports_ccli_preview(
    request: Request,
    start_date: str = Form(...),
    end_date: str = Form(...),
    db: Database = Depends(get_db),  # noqa: B008
    _rl: None = Depends(_report_rate_limit),  # noqa: B008
) -> HTMLResponse:
    """Inline summary of the CCLI report before download (#411).

    Lets a church admin confirm the date range covers data before downloading a
    CSV (which previously could be silently empty).
    """
    _validate_date_range(start_date, end_date)
    events = db.query_copy_events(start_date, end_date)
    unique_songs = len({e["song_id"] for e in events})
    return templates.TemplateResponse(
        request,
        "ccli_preview.html",
        {
            "start_date": start_date,
            "end_date": end_date,
            "event_count": len(events),
            "unique_songs": unique_songs,
        },
    )


@app.post("/reports/ccli")
async def reports_ccli_csv(
    start_date: str = Form(...),
    end_date: str = Form(...),
    db: Database = Depends(get_db),  # noqa: B008
    _rl: None = Depends(_report_rate_limit),  # noqa: B008
) -> StreamingResponse:
    """Generate CCLI compliance report as a CSV download (#201)."""
    _validate_date_range(start_date, end_date)
    events = db.query_copy_events(start_date, end_date)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Service", "Title", "CCLI#", "Reproduction Type", "Count"])
    for event in events:
        writer.writerow([
            event["service_date"],
            event["service_name"],
            event["display_title"],
            event.get("ccli_number", ""),
            event["reproduction_type"],
            event["count"],
        ])

    output.seek(0)
    filename = _sanitize_header_filename(f"ccli_report_{start_date}_{end_date}.csv")
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _missing_services_context(db: Database, days: int) -> dict[str, Any]:
    """Compute the report for *days* back from today, plus selector context."""
    data = compute_missing_services(db, days, date.today())
    data["window_options"] = WINDOW_OPTIONS
    return data


def _validate_exclusion_input(service_date: str, service_slot: str) -> None:
    """Raise HTTPException 422 for a bad exclusion date or slot."""
    if not _DATE_RE.match(service_date):
        raise HTTPException(
            status_code=422,
            detail=f"Invalid service_date: '{service_date}' — expected YYYY-MM-DD",
        )
    try:
        date.fromisoformat(service_date)
    except ValueError as exc:
        raise HTTPException(
            status_code=422,
            detail=f"Invalid service_date: '{service_date}' — not a real calendar date",
        ) from exc
    if service_slot not in VALID_SLOTS:
        raise HTTPException(
            status_code=422,
            detail=f"Invalid service_slot: '{service_slot}' — expected one of "
            f"{sorted(VALID_SLOTS)}",
        )


@app.get("/reports/missing-services", response_class=HTMLResponse)
async def missing_services_report(
    request: Request,
    days: int = Query(default=DEFAULT_WINDOW_DAYS),
    login: bool = Query(default=False),
    db: Database = Depends(get_db),  # noqa: B008
    _rl: None = Depends(_report_rate_limit),  # noqa: B008
) -> HTMLResponse:
    """Report Sunday morning/evening services absent from the DB (#480).

    The page + JSON are always publicly readable; the exclude/include controls
    render only for a viewer with valid upload credentials (#483). Visiting with
    ``?login=1`` issues the Basic-auth challenge so a user can authenticate, then
    the edit controls appear (the browser sends creds to this path thereafter).
    """
    authed = _upload_credentials_valid(request)
    if login and not authed:
        require_upload_auth(request)  # raises 401 challenge so the browser prompts
    ctx = _missing_services_context(db, days)
    ctx["can_edit"] = authed
    # HTMX swaps just the result region; a full navigation renders the page.
    template = (
        "_missing_services_result.html"
        if request.headers.get("HX-Request")
        else "missing_services.html"
    )
    return templates.TemplateResponse(request, template, ctx)


@app.get("/reports/missing-services.json")
async def missing_services_report_json(
    days: int = Query(default=DEFAULT_WINDOW_DAYS),
    db: Database = Depends(get_db),  # noqa: B008
    _rl: None = Depends(_report_rate_limit),  # noqa: B008
) -> JSONResponse:
    """JSON view of the missing-services report for automation/reuse (#480)."""
    return JSONResponse(content=_missing_services_context(db, days))


@app.post("/reports/missing-services/exclude", response_class=HTMLResponse)
async def missing_services_exclude(
    request: Request,
    service_date: str = Form(...),
    service_slot: str = Form(...),
    reason: str = Form(default=""),
    days: int = Form(default=DEFAULT_WINDOW_DAYS),
    db: Database = Depends(get_db),  # noqa: B008
    _rl: None = Depends(_report_rate_limit),  # noqa: B008
) -> HTMLResponse:
    """Mark a missing slot as intentionally excluded, then re-render (#480).

    Gated by the same upload auth that protects /upload (#483) — a write, not a
    read, so it must not be publicly invokable.
    """
    require_upload_auth(request)
    _validate_exclusion_input(service_date, service_slot)
    db.add_exclusion(service_date, service_slot, reason=reason or None)
    ctx = _missing_services_context(db, days)
    ctx["can_edit"] = True  # request is authenticated to have reached here
    return templates.TemplateResponse(request, "_missing_services_result.html", ctx)


@app.post("/reports/missing-services/include", response_class=HTMLResponse)
async def missing_services_include(
    request: Request,
    service_date: str = Form(...),
    service_slot: str = Form(...),
    days: int = Form(default=DEFAULT_WINDOW_DAYS),
    db: Database = Depends(get_db),  # noqa: B008
    _rl: None = Depends(_report_rate_limit),  # noqa: B008
) -> HTMLResponse:
    """Remove an intentional-exclusion marker, then re-render (#480).

    Gated by the same upload auth that protects /upload (#483).
    """
    require_upload_auth(request)
    _validate_exclusion_input(service_date, service_slot)
    db.remove_exclusion(service_date, service_slot)
    ctx = _missing_services_context(db, days)
    ctx["can_edit"] = True
    return templates.TemplateResponse(request, "_missing_services_result.html", ctx)


@app.get("/songs/{song_id}", response_class=HTMLResponse)
async def song_detail(
    request: Request,
    song_id: int,
    db: Database = Depends(get_db),  # noqa: B008
) -> HTMLResponse:
    song = db.query_song_by_id(song_id)
    if not song:
        _log.warning("Song not found", extra={"song_id": song_id})
        raise HTTPException(status_code=404, detail="Song not found")
    editions = db.query_song_editions(song_id)
    service_history = db.query_song_services(song_id)
    return templates.TemplateResponse(
        request,
        "song_detail.html",
        {"song": song, "editions": editions, "service_history": service_history},
    )


@app.get("/services", response_class=HTMLResponse)
async def services_list(
    request: Request,
    sort: str = Query(default="service_date"),
    sort_dir: str = Query(default="desc"),
    q_service: str = Query(default=""),
    q_leader: str = Query(default=""),
    q_preacher: str = Query(default=""),
    q_sermon: str = Query(default=""),
    start_date: str = Query(default=""),
    end_date: str = Query(default=""),
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=50, ge=10, le=500),
    db: Database = Depends(get_db),  # noqa: B008
) -> HTMLResponse:
    sort = sort if sort in Database._SERVICES_SORT_COLS else "service_date"
    sort_dir = "asc" if sort_dir == "asc" else "desc"
    services, total = db.query_all_services_paginated(
        sort=sort, sort_dir=sort_dir,
        q_service=q_service, q_leader=q_leader, q_preacher=q_preacher,
        q_sermon=q_sermon, start_date=start_date, end_date=end_date,
        page=page, per_page=per_page,
    )
    total_pages = math.ceil(total / per_page) if total > 0 else 1
    ctx = {
        "services": services, "sort": sort, "sort_dir": sort_dir,
        "q_service": q_service, "q_leader": q_leader, "q_preacher": q_preacher,
        "q_sermon": q_sermon, "start_date": start_date, "end_date": end_date,
        "page": page, "per_page": per_page, "total_pages": total_pages, "total": total,
    }
    if request.headers.get("HX-Request") and not _is_htmx_history_restore(request):
        return templates.TemplateResponse(request, "_services_results.html", ctx)
    ctx["service_names"] = db.query_distinct_service_names()
    ctx["leaders"] = db.query_distinct_song_leaders()
    ctx["preachers"] = db.query_distinct_preachers()
    return templates.TemplateResponse(request, "services.html", ctx)


@app.get("/services/{service_id}", response_class=HTMLResponse)
async def service_detail(
    request: Request,
    service_id: int,
    db: Database = Depends(get_db),  # noqa: B008
) -> HTMLResponse:
    service = db.query_service_by_id(service_id)
    if not service:
        _log.warning("Service not found", extra={"service_id": service_id})
        raise HTTPException(status_code=404, detail="Service not found")
    songs = db.query_service_songs(service_id)
    return templates.TemplateResponse(
        request, "service_detail.html", {"service": service, "songs": songs}
    )


_LEADER_MIN_SONG_COUNT = 2
_MIN_SERVICES_FOR_MEANINGFUL_TRENDS = 5


@app.get("/leaders", response_class=HTMLResponse)
async def leaders_index(
    request: Request,
    db: Database = Depends(get_db),  # noqa: B008
) -> HTMLResponse:
    leaders = db.query_all_leaders()
    return templates.TemplateResponse(request, "leaders.html", {"leaders": leaders})


@app.get("/leaders/{leader_name}/top-songs", response_class=HTMLResponse)
async def leader_top_songs(
    request: Request,
    leader_name: str,
    db: Database = Depends(get_db),  # noqa: B008
) -> HTMLResponse:
    top_songs = db.query_leader_top_songs(leader_name, min_count=_LEADER_MIN_SONG_COUNT)
    service_count = db.query_leader_service_count(leader_name)
    warning_few_services = service_count < _MIN_SERVICES_FOR_MEANINGFUL_TRENDS
    return templates.TemplateResponse(
        request,
        "leader_top_songs.html",
        {
            "leader": leader_name,
            "top_songs": top_songs,
            "service_count": service_count,
            "warning_few_services": warning_few_services,
            "min_count": _LEADER_MIN_SONG_COUNT,
        },
    )


@app.get("/leaders/{leader_name}/top-songs/csv")
async def leader_top_songs_csv(
    leader_name: str,
    db: Database = Depends(get_db),  # noqa: B008
    _rl: None = Depends(_report_rate_limit),  # noqa: B008
) -> StreamingResponse:
    top_songs = db.query_leader_top_songs(leader_name, min_count=_LEADER_MIN_SONG_COUNT)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Rank", "Title", "Credits", "Count"])
    for rank, song in enumerate(top_songs, 1):
        parts = [song.get("words_by") or "", song.get("music_by") or ""]
        credits = " / ".join(p for p in parts if p)
        writer.writerow([rank, song["display_title"], credits, song["performance_count"]])

    output.seek(0)
    safe_name = _sanitize_header_filename(leader_name)
    filename = f"leader_songs_{safe_name}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Upload + background import job endpoints (#45)
# ---------------------------------------------------------------------------


def _get_inbox_dir() -> Path:
    p = Path(os.environ.get("INBOX_DIR", "inbox"))
    p.mkdir(parents=True, exist_ok=True)
    return p


def _run_import_in_background(job_id: str, pptx_path: Path) -> None:
    """Import a PPTX file and update the job record when done.  Runs in a thread.

    All song/service DB writes are wrapped in a single transaction so that a
    mid-flight failure rolls back any partial inserts atomically.  The job
    status update is intentionally outside the transaction so it always commits
    even after a rollback.  The uploaded file is deleted in a finally block
    so that the inbox is always cleaned up regardless of success or failure.
    """
    db = _get_db()
    # Initialize notify variables with safe defaults so the finally block never
    # hits UnboundLocalError if db.update_import_job() raises (#193).
    _notify_title = "Import failed"
    _notify_message = f"{pptx_path.name} — unknown error"
    _notify_priority = -1
    try:
        result = run_import(db, pptx_path)

        if result.songs_imported == 0:
            db.update_import_job(
                job_id,
                status="complete",
                songs_imported=0,
                error_message="No songs found — file may not be a worship slide deck",
            )
            _notify_title = "Import complete (0 songs)"
            _notify_message = f"{pptx_path.name} — no songs found"
            _notify_priority = -1
        else:
            db.update_import_job(
                job_id,
                status="complete",
                songs_imported=result.songs_imported,
                service_date=result.service_date,
                service_name=result.service_name,
                song_leader=result.song_leader,
                preacher=result.preacher,
                sermon_title=result.sermon_title,
                songs_json=json.dumps(
                    [s.display_title for s in result.songs]
                ) if result.songs else None,
            )
            _notify_title = "Import complete"
            _notify_message = (
                f"{pptx_path.name} — {result.songs_imported} songs imported"
            )
            _notify_priority = 0
    except Exception as exc:  # noqa: BLE001
        # update_import_job runs outside the transaction block — commits even on rollback
        db.update_import_job(
            job_id, status="failed", error_message=str(exc)[:500]
        )
        _notify_title = "Import failed"
        _notify_message = f"{pptx_path.name} — {exc}"
        _notify_priority = -1
    finally:
        db.close()
        # Notification is fire-and-forget — isolated from job status updates (#185)
        try:
            send_pushover(
                title=_notify_title,
                message=_notify_message,
                priority=_notify_priority,
            )
        except Exception:  # noqa: BLE001
            _log.warning("Pushover notification could not be sent", exc_info=True)
        # Always clean up the inbox file regardless of success or failure (#138)
        try:
            if pptx_path.exists():
                pptx_path.unlink()
        except OSError as exc:  # noqa: BLE001
            _log.warning(
                "Failed to delete uploaded file from inbox",
                extra={"path": str(pptx_path), "error": str(exc)},
            )


# Default username for the upload Basic Auth gate; the password is the secret.
_UPLOAD_USERNAME_DEFAULT = "highland"


def _upload_credentials_valid(request: Request) -> bool:
    """True when upload auth is satisfied for this request (non-raising).

    Returns True when ``UPLOAD_PASSWORD`` is unset (auth disabled — open access,
    current behavior) or when the request carries matching HTTP Basic
    credentials. This is the companion to :func:`require_upload_auth`, used to
    decide whether to render edit controls on an otherwise-public page (#483).
    """
    password = os.environ.get("UPLOAD_PASSWORD")
    if not password:
        return True
    expected_user = os.environ.get("UPLOAD_USERNAME", _UPLOAD_USERNAME_DEFAULT)
    header = request.headers.get("Authorization", "")
    scheme, _, encoded = header.partition(" ")
    if scheme.lower() != "basic" or not encoded:
        return False
    try:
        decoded = base64.b64decode(encoded).decode("utf-8")
    except (ValueError, UnicodeDecodeError):
        return False
    user, _, supplied = decoded.partition(":")
    return secrets.compare_digest(user, expected_user) and secrets.compare_digest(
        supplied, password
    )


def require_upload_auth(request: Request) -> None:
    """Gate write routes behind HTTP Basic Auth (#388, #483).

    Active only when ``UPLOAD_PASSWORD`` is set in the environment; when unset
    the write stays open (current behavior). The catalog browsing/report pages
    are never gated — only write access (upload + exclusion edits). The realm is
    shared so a single login unlocks every write surface.
    """
    if _upload_credentials_valid(request):
        return
    raise HTTPException(
        status_code=401,
        detail="Authentication required to upload.",
        headers={"WWW-Authenticate": 'Basic realm="Highland Worship Catalog upload"'},
    )


@app.get("/upload", response_class=HTMLResponse)
async def upload_page(
    request: Request,
    login: bool = Query(default=False),
) -> HTMLResponse:
    """Render the browser upload form for PPTX files.

    Mirrors the missing-services report (#483, #485): the page is always
    viewable, but the upload control renders only for a viewer with valid upload
    credentials — otherwise an inert "Log in to upload" state is shown. Visiting
    with ``?login=1`` issues the Basic-auth challenge so the browser prompts and
    caches creds for the /upload realm. POST /upload stays auth-gated, so the
    inert front-end is never the only guard.
    """
    authed = _upload_credentials_valid(request)
    if login and not authed:
        require_upload_auth(request)  # raises 401 challenge so the browser prompts
    return templates.TemplateResponse(
        request,
        "upload.html",
        {"authed": authed, "auth_required": bool(os.environ.get("UPLOAD_PASSWORD"))},
    )


@app.post("/upload")
async def upload(
    request: Request,
    file: UploadFile,
    db: Database = Depends(get_db),  # noqa: B008
    _: None = Depends(require_upload_auth),  # noqa: B008
) -> JSONResponse:
    """Accept a PPTX file, create an import job, and kick off background import."""
    # Rate limiting (#173) — check before reading the body to save bandwidth.
    # Prefer X-Forwarded-For when behind a reverse proxy (#283).
    client_ip = _get_client_ip(request)
    allowed, retry_after = _upload_limiter.is_allowed(client_ip)
    if not allowed:
        _log.warning(
            "Upload rate limit exceeded",
            extra={"client_ip": client_ip, "retry_after": retry_after},
        )
        return JSONResponse(
            content={"detail": "Upload rate limit exceeded — try again later"},
            status_code=429,
            headers={"Retry-After": str(retry_after)},
        )
    # Pre-flight: reject by Content-Length header before reading the body (defence-in-depth)
    cl_header = request.headers.get("content-length")
    if cl_header is not None:
        try:
            if int(cl_header) > MAX_UPLOAD_BYTES:
                _max_mb = MAX_UPLOAD_BYTES // (1024 * 1024)
                send_pushover(
                    title="Upload rejected",
                    message=f"File exceeds {_max_mb} MB limit (Content-Length: {cl_header})",
                    priority=-1,
                )
                return JSONResponse(
                    content={"detail": f"File exceeds maximum allowed size of {MAX_UPLOAD_BYTES // (1024 * 1024)} MB"},  # noqa: E501
                    status_code=413,
                )
        except ValueError:
            return JSONResponse(
                content={"detail": "Invalid Content-Length header"}, status_code=400
            )
    # Validate MIME type
    if file.content_type != _PPTX_MIME:
        send_pushover(
            title="Upload rejected",
            message=f"Wrong MIME type: {file.content_type}",
            priority=-1,
        )
        return JSONResponse(
            content={"detail": "Only PPTX files are accepted (pptx mime type required)"},
            status_code=400,
        )
    # Sanitize filename: strip directory components and unsafe characters (#106)
    raw_filename = file.filename or ""
    filename = _sanitize_header_filename(raw_filename)
    # Validate extension (re-check after sanitization)
    if not filename.lower().endswith(".pptx"):
        send_pushover(
            title="Upload rejected",
            message=f"{filename} — file must have a .pptx extension",
            priority=-1,
        )
        return JSONResponse(
            content={"detail": "File must have a .pptx extension"},
            status_code=400,
        )
    # Validate that something remains after sanitization
    stem = filename[: -len(".pptx")]
    if not stem:
        return JSONResponse(
            content={"detail": "Filename is invalid after sanitization"},
            status_code=400,
        )
    # Read body in chunks to bound memory usage (#297)
    chunks: list[bytes] = []
    total_read = 0
    while True:
        chunk = await file.read(_UPLOAD_CHUNK_SIZE)
        if not chunk:
            break
        total_read += len(chunk)
        if total_read > MAX_UPLOAD_BYTES:
            break
        chunks.append(chunk)
    content = b"".join(chunks)
    # Validate ZIP magic bytes — PPTX is a ZIP archive (PK\x03\x04) (#320)
    if len(content) < 4 or content[:4] != b"PK\x03\x04":
        send_pushover(
            title="Upload rejected",
            message=f"{filename} — not a valid PPTX (bad magic bytes)",
            priority=-1,
        )
        return JSONResponse(
            content={"detail": "File is not a valid PPTX archive"},
            status_code=400,
        )
    if total_read > MAX_UPLOAD_BYTES:
        limit_mb = MAX_UPLOAD_BYTES // (1024 * 1024)
        send_pushover(
            title="Upload rejected",
            message=f"{filename} — file exceeds {limit_mb} MB limit",
            priority=-1,
        )
        return JSONResponse(
            content={"detail": f"File exceeds maximum allowed size of {limit_mb} MB"},
            status_code=413,
        )
    # Save to inbox
    inbox = _get_inbox_dir()
    job_id = secrets.token_urlsafe(32)
    dest = inbox / f"{job_id}_{filename}"
    dest.write_bytes(content)
    # Create pending job record
    db.create_import_job(job_id, filename=filename)
    # Submit import to bounded thread pool (#52).
    # If the pool is saturated or shut down, submit() raises — return 503.
    try:
        _import_executor.submit(_run_import_in_background, job_id, dest)
    except Exception as exc:  # noqa: BLE001
        _log.warning("Import pool unavailable, rejecting upload", extra={"error": str(exc)})
        return JSONResponse(
            content={"detail": "Server busy — import queue is full, try again later"},
            status_code=503,
        )
    return JSONResponse(content={"job_id": job_id}, status_code=202)


@app.get("/jobs")
async def list_jobs(
    db: Database = Depends(get_db),  # noqa: B008
    _: None = Depends(require_upload_auth),  # noqa: B008
) -> JSONResponse:
    """Return all import job records, newest first.

    Gated behind the upload auth (#451): the job log leaks uploaded filenames
    and raw error messages, so it's part of the write/admin surface, not public.
    """
    jobs = db.list_import_jobs()
    return JSONResponse(content=jobs)


@app.get("/jobs/{job_id}")
async def get_job(
    job_id: str,
    db: Database = Depends(get_db),  # noqa: B008
    _: None = Depends(require_upload_auth),  # noqa: B008
) -> JSONResponse:
    """Return a single import job record or 404 (auth-gated like /jobs, #451)."""
    job = db.get_import_job(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Job not found")
    return JSONResponse(content=job)
